import pandas as pd
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


class PersonalTranscriptGenerator:
    def __init__(self):
        self.source_file = (
            r"25年9月体测成绩得分等级汇总/全校学生体质健康测试成绩总表.xlsx"
        )
        self.output_dir = r"25年9月体测成绩得分等级汇总\成绩表"
        self.today = "2025/09/23-2025/09/24"

        # 跳绳满分标准 (年级: {性别: 满分次数})
        self.jump_rope_standards = {
            "一年级": {"男": 109, "女": 117},
            "二年级": {"男": 117, "女": 127},
            "三年级": {"男": 126, "女": 139},
            "四年级": {"男": 137, "女": 149},
            "五年级": {"男": 148, "女": 158},
            "六年级": {"男": 157, "女": 166},
        }

        # 确保输出目录存在
        os.makedirs(self.output_dir, exist_ok=True)

    def load_student_data(self):
        """加载学生数据"""
        print("=== 加载全校学生数据 ===")

        try:
            # 先读取整个文件看看结构
            df_all = pd.read_excel(self.source_file, header=None)
            print(f"Excel文件总行数: {len(df_all)}")

            # 读取Excel文件，跳过前7行（统计信息），第8行是表头
            df = pd.read_excel(self.source_file, skiprows=7)
            print(f"跳过前7行后的数据行数: {len(df)}")

            # 直接使用Excel的列名，不重新命名
            print(f"实际列名: {df.columns.tolist()}")

            # 创建列名映射（基于实际的列名结构）
            self.column_mapping = {
                "年级编号": "年级",
                "班名": "班级",
                "学号": "学号",
                "姓名": "姓名",
                "性别": "性别",
                "身高（cm)": "身高",
                "体重（kg)": "体重",
                "体重指数BMI\n（千克/米2）": "BMI值",
                "得分": "BMI得分",
                "等级": "BMI等级",
                "肺活量（毫升）": "肺活量值",
                "得分.1": "肺活量得分",
                "等级.1": "肺活量等级",
                "50米跑（秒）": "50米跑值",
                "得分.2": "50米跑得分",
                "等级.2": "50米跑等级",
                "坐位体前屈(cm)": "坐位体前屈值",
                "得分.3": "坐位体前屈得分",
                "等级.3": "坐位体前屈等级",
                "一分钟跳绳（次）": "跳绳值",
                "得分.4": "跳绳得分",
                "加分": "跳绳加分",  # 新增的加分列
                "等级.4": "跳绳等级",
                "一分钟仰卧起坐（次）": "仰卧起坐值",
                "得分.5": "仰卧起坐得分",
                "等级.5": "仰卧起坐等级",
                "50米*8往返跑（分.秒）": "50米×8往返跑值",
                "得分.6": "50米×8往返跑得分",
                "等级.6": "50米×8往返跑等级",
                "标准分": "标准分",
                "附加分": "附加分",
                "综合得分": "综合得分",
                "综合评级": "综合等级",
            }

            # 重命名列
            df = df.rename(columns=self.column_mapping)

            # 检查空数据情况
            print(f"学号为空的行数: {df['学号'].isna().sum()}")
            print(f"姓名为空的行数: {df['姓名'].isna().sum()}")

            # 查看所有非空的学号和姓名组合
            valid_students = df[df["学号"].notna() & df["姓名"].notna()]
            print(f"有效学号和姓名的学生数: {len(valid_students)}")

            # 不进行清理，保留所有数据，但在生成时跳过无效数据
            print(f"保留所有数据 {len(df)} 行")
            return df

        except Exception as e:
            print(f"加载数据时出错: {e}")
            return pd.DataFrame()

    def calculate_jump_rope_display(self, grade, gender, actual_jumps, jump_score):
        """计算跳绳显示内容和加分"""
        if not grade or not actual_jumps or not gender:
            return actual_jumps, jump_score, "", 0, 0

        try:
            actual_jumps = int(actual_jumps)
            jump_score = int(jump_score) if jump_score else 0

            # 获取该年级性别的满分标准
            grade_standards = self.jump_rope_standards.get(grade, {})
            standard_jumps = grade_standards.get(gender, 0)

            # 根据分数确定等级
            if jump_score >= 90:
                grade_level = "优秀"
            elif jump_score >= 80:
                grade_level = "良好"
            elif jump_score >= 60:
                grade_level = "及格"
            else:
                grade_level = "不及格"

            if jump_score == 100:
                # 得分是100分：计算加分
                if actual_jumps > standard_jumps:
                    # 实际超过满分，计算超出个数和加分
                    extra_jumps = actual_jumps - standard_jumps
                    bonus_score = min(extra_jumps // 2, 20)
                    return (
                        actual_jumps,
                        jump_score,
                        grade_level,
                        extra_jumps,
                        bonus_score,
                    )
                else:
                    # 实际没超过满分，无加分
                    return actual_jumps, jump_score, grade_level, 0, 0
            else:
                # 得分不是100分：无加分
                return actual_jumps, jump_score, grade_level, 0, 0

        except:
            return actual_jumps, jump_score, "", 0, 0

    def create_personal_transcript(self, student_data):
        """为单个学生创建个人成绩表"""
        wb = Workbook()
        ws = wb.active
        ws.title = "成绩单"

        # 设置列宽
        col_widths = [
            28,
            10,
            13,
            10,
            23,
            7,
            9,
        ]  # A=28, B=10, C=13, D=10, E=23, F=7, G=9
        for i, width in enumerate(col_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = width

        # 创建边框样式
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 标题
        ws.merge_cells("A1:G1")
        ws["A1"] = "《国家学生体质健康标准》登记卡"
        ws["A1"].font = Font(bold=True, size=14)

        # 第2行：学校和测评日期
        ws["A2"] = "学    校"
        ws["A2"].border = thin_border
        ws.merge_cells("B2:C2")
        ws["B2"] = "平谷区第十一小学"
        ws["B2"].border = thin_border
        ws["D2"] = "测评日期"
        ws["D2"].border = thin_border
        ws.merge_cells("E2:G2")
        # 统一设置测评日期为9.23-9.24
        ws["E2"] = "2025/09/23-2025/09/24"
        ws["E2"].border = thin_border

        # 第3行：学生基本信息
        ws["A3"] = "姓    名"
        ws["A3"].border = thin_border
        ws["B3"] = student_data.get("姓名", "")
        ws["B3"].border = thin_border
        ws["C3"] = "性  别"
        ws["C3"].border = thin_border
        ws["D3"] = student_data.get("性别", "")
        ws["D3"].border = thin_border
        ws["E3"] = "学    号"
        ws["E3"].border = thin_border
        ws.merge_cells("F3:G3")
        ws["F3"] = student_data.get("学号", "")
        ws["F3"].border = thin_border

        # 第4行：班级、民族
        ws["A4"] = "班    级"
        ws["A4"].border = thin_border
        ws.merge_cells("B4:C4")
        ws["B4"] = f"{student_data.get('年级', '')}{student_data.get('班级', '')}"
        ws["B4"].border = thin_border
        ws["D4"] = "民  族"
        ws["D4"].border = thin_border
        ws.merge_cells("E4:G4")
        ws["E4"] = "汉族"
        ws["E4"].border = thin_border

        # 第5行：身高、体重
        ws["A5"] = "身高（cm）"
        ws["A5"].border = thin_border
        ws["B5"] = student_data.get("身高", "")
        ws["B5"].border = thin_border
        ws["C5"] = "体重（kg）"
        ws["C5"].border = thin_border
        ws["D5"] = student_data.get("体重", "")
        ws["D5"].border = thin_border
        # 空白区域
        ws.merge_cells("E5:G5")
        ws["E5"] = ""
        ws["E5"].border = thin_border

        # 第6行：表头
        ws["A6"] = "单项指标"
        ws["A6"].border = thin_border
        ws["B6"] = "成绩"
        ws["B6"].border = thin_border
        ws["C6"] = "得分"
        ws["C6"].border = thin_border
        ws["D6"] = "等级"
        ws["D6"].border = thin_border
        ws["E6"] = "加分指标"
        ws["E6"].border = thin_border
        ws["F6"] = "成绩"
        ws["F6"].border = thin_border
        ws["G6"] = "附加分"
        ws["G6"].border = thin_border

        # 设置表头居中和字体
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{col}6"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"{col}6"].font = Font(bold=True, size=11)

        # 第7行：BMI
        ws["A7"] = "体重指数（BMI）\n（千克/米²）"
        ws["A7"].border = thin_border
        ws["A7"].alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws["B7"] = student_data.get("BMI值", "")
        ws["B7"].border = thin_border
        ws["B7"].alignment = Alignment(horizontal="center", vertical="center")
        ws["C7"] = student_data.get("BMI得分", "")
        ws["C7"].border = thin_border
        ws["C7"].alignment = Alignment(horizontal="center", vertical="center")
        ws["D7"] = student_data.get("BMI等级", "")
        ws["D7"].border = thin_border
        ws["D7"].alignment = Alignment(horizontal="center", vertical="center")

        # 获取跳绳数据和加分
        grade = student_data.get("年级", "")
        gender = student_data.get("性别", "")
        actual_jumps = student_data.get("跳绳值", "")
        jump_score = student_data.get("跳绳得分", "")
        jump_bonus = student_data.get("跳绳加分", 0)  # 直接从数据中获取加分
        jump_grade = student_data.get("跳绳等级", "")

        # 如果有加分，计算超出次数
        extra_jumps = 0
        if jump_bonus and pd.notna(jump_bonus) and jump_bonus > 0:
            extra_jumps = int(jump_bonus * 2)  # 反推超出次数

        ws["E7"] = "1 分钟跳绳\n（单位：次）"
        ws["E7"].border = thin_border
        ws["E7"].alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        # F7显示超出满分的个数，G7显示加分
        if extra_jumps > 0:
            ws["F7"] = extra_jumps
            ws["G7"] = jump_bonus
        else:
            ws["F7"] = ""
            ws["G7"] = ""
        ws["F7"].border = thin_border
        ws["F7"].alignment = Alignment(horizontal="center", vertical="center")
        ws["G7"].border = thin_border
        ws["G7"].alignment = Alignment(horizontal="center", vertical="center")

        # 第8行：肺活量
        ws["A8"] = "肺活量（毫升）"
        ws["A8"].border = thin_border
        ws["A8"].alignment = Alignment(horizontal="center", vertical="center")
        ws["B8"] = student_data.get("肺活量值", "")
        ws["B8"].border = thin_border
        ws["B8"].alignment = Alignment(horizontal="center", vertical="center")
        ws["C8"] = student_data.get("肺活量得分", "")
        ws["C8"].border = thin_border
        ws["C8"].alignment = Alignment(horizontal="center", vertical="center")
        ws["D8"] = student_data.get("肺活量等级", "")
        ws["D8"].border = thin_border
        ws["D8"].alignment = Alignment(horizontal="center", vertical="center")

        # 学年总分
        ws["E8"] = "学年总分"
        ws["E8"].border = thin_border
        ws["E8"].alignment = Alignment(horizontal="center", vertical="center")
        standard_score = student_data.get("标准分", 0)
        additional_score = student_data.get("附加分", 0)
        total_score = (standard_score if pd.notna(standard_score) else 0) + (
            additional_score if pd.notna(additional_score) else 0
        )
        ws.merge_cells("F8:G8")
        ws["F8"] = f"{total_score:.1f}" if total_score > 0 else ""
        ws["F8"].border = thin_border
        ws["F8"].alignment = Alignment(horizontal="center")

        # 第9行：50米跑
        ws["A9"] = "50 米跑（秒）"
        ws["A9"].border = thin_border
        ws["A9"].alignment = Alignment(horizontal="center", vertical="center")
        ws["B9"] = student_data.get("50米跑值", "")
        ws["B9"].border = thin_border
        ws["C9"] = student_data.get("50米跑得分", "")
        ws["C9"].border = thin_border
        ws["D9"] = student_data.get("50米跑等级", "")
        ws["D9"].border = thin_border
        ws["E9"] = ""
        ws["E9"].border = thin_border
        ws.merge_cells("F9:G9")
        ws["F9"] = ""
        ws["F9"].border = thin_border

        # 第10行：坐位体前屈
        ws["A10"] = "坐位体前屈\n（厘米）"
        ws["A10"].border = thin_border
        ws["A10"].alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws["B10"] = student_data.get("坐位体前屈值", "")
        ws["B10"].border = thin_border
        ws["C10"] = student_data.get("坐位体前屈得分", "")
        ws["C10"].border = thin_border
        ws["D10"] = student_data.get("坐位体前屈等级", "")
        ws["D10"].border = thin_border

        # 等级评定
        ws["E10"] = "等级评定"
        ws["E10"].border = thin_border
        ws["E10"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("F10:G10")
        ws["F10"] = student_data.get("综合等级", "")
        ws["F10"].border = thin_border
        ws["F10"].alignment = Alignment(horizontal="center")

        # 第11行：跳绳（计算后的显示结果）
        ws["A11"] = "1 分钟跳绳\n（单位：次）"
        ws["A11"].border = thin_border
        ws["A11"].alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws["B11"] = actual_jumps if actual_jumps and pd.notna(actual_jumps) else ""
        ws["B11"].border = thin_border
        ws["C11"] = jump_score if jump_score and pd.notna(jump_score) else ""
        ws["C11"].border = thin_border
        ws["D11"] = jump_grade if jump_grade and pd.notna(jump_grade) else ""
        ws["D11"].border = thin_border
        ws["E11"] = ""
        ws["E11"].border = thin_border
        ws.merge_cells("F11:G11")
        ws["F11"] = ""
        ws["F11"].border = thin_border

        # 第12行：仰卧起坐
        ws["A12"] = "1 分钟仰卧起坐\n（单位：次）"
        ws["A12"].border = thin_border
        ws["A12"].alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws["B12"] = (
            student_data.get("仰卧起坐值", "")
            if pd.notna(student_data.get("仰卧起坐值", ""))
            else ""
        )
        ws["B12"].border = thin_border
        ws["C12"] = (
            student_data.get("仰卧起坐得分", "")
            if pd.notna(student_data.get("仰卧起坐得分", ""))
            else ""
        )
        ws["C12"].border = thin_border
        ws["D12"] = (
            student_data.get("仰卧起坐等级", "")
            if pd.notna(student_data.get("仰卧起坐等级", ""))
            else ""
        )
        ws["D12"].border = thin_border

        # 体育教师签字
        ws["E12"] = "体育教师签字"
        ws["E12"].border = thin_border
        ws["E12"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("F12:G12")
        ws["F12"] = ""
        ws["F12"].border = thin_border

        # 第13行：50米×8往返跑（仅六年级需要）
        if student_data.get("年级", "") == "六年级":
            ws["A13"] = "50米×8往返跑\n（单位：s）"
            ws["A13"].border = thin_border
            ws["A13"].alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            ws["B13"] = (
                student_data.get("50米×8往返跑值", "")
                if pd.notna(student_data.get("50米×8往返跑值", ""))
                else ""
            )
            ws["B13"].border = thin_border
            ws["C13"] = (
                student_data.get("50米×8往返跑得分", "")
                if pd.notna(student_data.get("50米×8往返跑得分", ""))
                else ""
            )
            ws["C13"].border = thin_border
            ws["D13"] = (
                student_data.get("50米×8往返跑等级", "")
                if pd.notna(student_data.get("50米×8往返跑等级", ""))
                else ""
            )
            ws["D13"].border = thin_border
            ws["E13"] = ""
            ws["E13"].border = thin_border
            ws.merge_cells("F13:G13")
            ws["F13"] = ""
            ws["F13"].border = thin_border
        else:
            # 四年级：空行，仅保持边框
            for col in ["A", "B", "C", "D", "E"]:
                ws[f"{col}13"] = ""
                ws[f"{col}13"].border = thin_border
            ws.merge_cells("F13:G13")
            ws["F13"] = ""
            ws["F13"].border = thin_border

        # 第14行：班主任签字
        ws["E14"] = "班主任签字"
        ws["E14"].border = thin_border
        ws["E14"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("F14:G14")
        ws["F14"] = ""
        ws["F14"].border = thin_border
        for col in ["A", "B", "C", "D"]:
            ws[f"{col}14"].border = thin_border

        # 第15行：家长签字
        ws["E15"] = "家长签字"
        ws["E15"].border = thin_border
        ws["E15"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("F15:G15")
        ws["F15"] = ""
        ws["F15"].border = thin_border
        for col in ["A", "B", "C", "D"]:
            ws[f"{col}15"].border = thin_border

        # 第16行：空行
        for col in ["A", "B", "C", "D", "E"]:
            ws[f"{col}16"].border = thin_border
        ws.merge_cells("F16:G16")
        ws["F16"].border = thin_border

        # 第17行：标准分和附加分
        ws["A17"] = "标准分"
        ws["A17"].border = thin_border
        ws["A17"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("B17:C17")
        ws["B17"] = (
            f"{student_data.get('标准分', ''):.1f}"
            if pd.notna(student_data.get("标准分", ""))
            else ""
        )
        ws["B17"].border = thin_border
        ws["B17"].alignment = Alignment(horizontal="center")

        ws["D17"] = "附加分"
        ws["D17"].border = thin_border
        ws["D17"].alignment = Alignment(horizontal="center", vertical="center")
        ws["E17"] = (
            additional_score
            if additional_score and pd.notna(additional_score) and additional_score > 0
            else ""
        )
        ws["E17"].border = thin_border
        ws["E17"].alignment = Alignment(horizontal="center")

        ws["F17"] = "备注"
        ws["F17"].border = thin_border
        ws["F17"].alignment = Alignment(horizontal="center", vertical="center")
        ws["G17"] = ""
        ws["G17"].border = thin_border

        # 确保所有单元格都有边框和居中对齐
        for row in range(1, 18):
            for col in ["A", "B", "C", "D", "E", "F", "G"]:
                cell = ws[f"{col}{row}"]
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # 设置行高
        for row_num in range(1, 18):
            ws.row_dimensions[row_num].height = 30

        return wb

    def generate_all_transcripts(self):
        """为所有学生生成个人成绩表，按年级和班级分文件夹保存"""
        print("=== 开始生成个人成绩表 ===")
        print("只处理四年级和六年级的数据")

        # 加载学生数据
        students_df = self.load_student_data()

        if students_df.empty:
            print("没有找到学生数据，退出生成")
            return

        # 过滤只保留四年级和六年级的学生
        target_grades = ["四年级", "六年级"]
        students_df = students_df[students_df["年级"].isin(target_grades)]
        print(f"过滤后剩余 {len(students_df)} 条四年级和六年级学生数据")

        success_count = 0
        total_count = len(students_df)

        for idx, student in students_df.iterrows():
            try:
                student_id = str(student.get("学号", "")).strip()
                student_name = str(student.get("姓名", "")).strip()
                grade = str(student.get("年级", "")).strip()
                class_name = str(student.get("班级", "")).strip()

                if not student_name or student_name == "nan":
                    print(f"跳过无效学生数据(行{idx+1}): 姓名={student_name}")
                    continue

                # 创建年级文件夹路径
                grade_dir = (
                    os.path.join(self.output_dir, grade) if grade else self.output_dir
                )

                # 创建班级文件夹路径
                class_dir = (
                    os.path.join(grade_dir, class_name) if class_name else grade_dir
                )

                # 确保文件夹存在
                os.makedirs(class_dir, exist_ok=True)

                # 创建个人成绩表
                wb = self.create_personal_transcript(student)

                # 生成文件名（优先使用学号）
                if student_id and student_id != "nan":
                    filename = f"{student_id}_{student_name}_成绩单.xlsx"
                else:
                    filename = f"{student_name}_成绩单.xlsx"
                filepath = os.path.join(class_dir, filename)

                # 保存文件
                wb.save(filepath)
                success_count += 1

                if success_count % 50 == 0:
                    print(f"已生成 {success_count}/{total_count} 个成绩表")

            except Exception as e:
                print(f"生成学生 {student.get('姓名', 'Unknown')} 成绩表时出错: {e}")

        print(f"\n=== 生成完成 ===")
        print(f"成功生成 {success_count}/{total_count} 个个人成绩表")
        print(f"保存位置: {self.output_dir}")
        print(f"文件已按年级和班级分文件夹保存")


def main():
    generator = PersonalTranscriptGenerator()
    generator.generate_all_transcripts()


if __name__ == "__main__":
    main()
