import pandas as pd
import numpy as np
import os
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime

class ClassRankingGenerator:
    def __init__(self):
        self.input_file = "2025年09月23日-2025年09月24日 成绩_含学号.xlsx"
        self.output_dir = r'25年9月体测成绩得分等级汇总'

        self.class_summary_dir = r'25年9月体测成绩得分等级汇总' 
        self.today = datetime.now().strftime("%Y/%m/%d")
        
        # 确保输出目录存在
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
        
        self.output_file = os.path.join(self.output_dir, '班级排名统计表.xlsx')
    
    def extract_grade_from_class(self, class_name):
        """从班级名称中提取年级"""
        if pd.isna(class_name):
            return None
        
        class_str = str(class_name)
        for grade in ['一年级', '二年级', '三年级', '四年级', '五年级', '六年级']:
            if grade in class_str:
                return grade
        return None
    
    def load_grade_data(self):
        """加载Excel文件，获取班级基本信息"""
        grade_class_info = {}
        
        print("=== 加载学生数据获取班级信息 ===")
        print(f"读取文件: {self.input_file}")
        print("只处理四年级和六年级的数据")
        
        try:
            df = pd.read_excel(self.input_file)
            print(f"读取到 {len(df)} 条数据")
            
            # 列名映射以适应新的Excel格式
            column_mapping = {
                '学籍号': '学号'
            }
            
            # 应用列名映射
            df = df.rename(columns=column_mapping)
            
            # 数据清理（只检查姓名）
            df_clean = df.dropna(subset=['姓名'])
            print(f"清理后剩余 {len(df_clean)} 条有效数据")
            
            # 添加年级信息
            df_clean['年级'] = df_clean['班级名称'].apply(self.extract_grade_from_class)
            
            # 过滤只保留四年级和六年级
            target_grades = ['四年级', '六年级']
            df_clean = df_clean[df_clean['年级'].isin(target_grades)]
            print(f"过滤后剩余 {len(df_clean)} 条四年级和六年级数据")
            
            # 按年级分组统计班级信息
            for grade in df_clean['年级'].dropna().unique():
                grade_data = df_clean[df_clean['年级'] == grade]
                class_stats = {}
                
                for class_name in grade_data['班级名称'].unique():
                    if pd.isna(class_name):
                        continue
                    
                    class_df = grade_data[grade_data['班级名称'] == class_name]
                    
                    # 提取班号
                    import re
                    match = re.search(r'(\d+)班', str(class_name))
                    class_num = int(match.group(1)) if match else 999
                    
                    # 格式化班级名
                    grade_map = {'一年级': '一', '二年级': '二', '三年级': '三', 
                               '四年级': '四', '五年级': '五', '六年级': '六'}
                    grade_short = grade_map.get(grade, grade)
                    formatted_class_name = f"{grade_short}{class_num}班"
                    
                    # 统计基本信息 - 使用更严格的测试数据检查
                    total_students = len(class_df)
                    # 有任何一项测试数据就算参与测试
                    test_columns = ['身高(cm)', '体重(kg)', '肺活量(ml)', '50米跑(s)', '坐位体前屈(cm)', '一分钟跳绳(个）']
                    available_columns = [col for col in test_columns if col in class_df.columns]
                    
                    if available_columns:
                        tested_students = len(class_df.dropna(subset=available_columns, how='all'))
                    else:
                        tested_students = 0
                    
                    class_stats[formatted_class_name] = {
                        'class_num': class_num,
                        'total_students': total_students,
                        'tested_students': tested_students,
                        'test_rate': tested_students / total_students * 100 if total_students > 0 else 0
                    }
                
                if class_stats:  # 只添加有班级数据的年级
                    grade_class_info[grade] = class_stats
                    print(f"{grade}: {len(class_stats)} 个班级")
            
        except Exception as e:
            print(f"读取文件 {self.input_file} 时出错: {e}")
        
        return grade_class_info
    
    def load_class_summary_data(self, grade):
        """读取班级汇总表数据"""
        grade_dir = os.path.join(self.class_summary_dir, grade)
        if not os.path.exists(grade_dir):
            print(f"年级目录不存在: {grade_dir}")
            return {}
        
        class_summary_data = {}
        
        # 查找所有班级汇总表 - 修复文件名模式匹配
        # 将"四年级"转换为"四"来匹配实际文件名格式
        grade_map = {'一年级': '一', '二年级': '二', '三年级': '三', 
                     '四年级': '四', '五年级': '五', '六年级': '六'}
        grade_short = grade_map.get(grade, grade)
        pattern = os.path.join(grade_dir, f"{grade_short}*班_班级统计汇总表.xlsx")
        summary_files = glob.glob(pattern)
        
        print(f"在{grade}目录中找到 {len(summary_files)} 个班级汇总表")
        
        for file_path in summary_files:
            try:
                # 从文件名提取班级信息
                filename = os.path.basename(file_path)
                import re
                # 修复正则表达式匹配实际文件名格式：四1班_班级统计汇总表.xlsx
                match = re.search(r'([一二三四五六]\d+班)_班级统计汇总表\.xlsx', filename)
                if not match:
                    continue
                
                class_name = match.group(1)  # 如: 四10班
                
                # 读取Excel文件
                df = pd.read_excel(file_path, header=None)
                
                # 提取关键数据
                summary_info = {}
                
                # 查找应查人数行
                for idx, row in df.iterrows():
                    if pd.notna(row.iloc[0]) and '应查人数' in str(row.iloc[0]):
                        summary_info['total_students'] = row.iloc[6] if pd.notna(row.iloc[6]) else 0
                        break
                
                # 查找实查人数行
                for idx, row in df.iterrows():
                    if pd.notna(row.iloc[0]) and '实查人数' in str(row.iloc[0]):
                        summary_info['tested_students'] = row.iloc[6] if pd.notna(row.iloc[6]) else 0
                        break
                
                # 查找实查比率行
                for idx, row in df.iterrows():
                    if pd.notna(row.iloc[0]) and '实查比率' in str(row.iloc[0]):
                        rate_str = str(row.iloc[6]) if pd.notna(row.iloc[6]) else "0%"
                        summary_info['test_rate'] = float(rate_str.replace('%', '')) if '%' in rate_str else 0
                        break
                
                # 查找各等级人数和占比
                for idx, row in df.iterrows():
                    if pd.notna(row.iloc[7]) and '一级（优秀）' in str(row.iloc[7]):
                        summary_info['excellent_count'] = row.iloc[16] if pd.notna(row.iloc[16]) else 0
                        rate_str = str(row.iloc[17]) if pd.notna(row.iloc[17]) else "0%"
                        summary_info['excellent_rate'] = float(rate_str.replace('%', '')) if '%' in rate_str else 0
                    elif pd.notna(row.iloc[7]) and '二级（良好）' in str(row.iloc[7]):
                        summary_info['good_count'] = row.iloc[16] if pd.notna(row.iloc[16]) else 0
                    elif pd.notna(row.iloc[7]) and '三级（及格）' in str(row.iloc[7]):
                        summary_info['pass_count'] = row.iloc[16] if pd.notna(row.iloc[16]) else 0
                    elif pd.notna(row.iloc[7]) and '四级（不及格）' in str(row.iloc[7]):
                        summary_info['fail_count'] = row.iloc[16] if pd.notna(row.iloc[16]) else 0
                
                # 计算平均分（基于等级分布估算）
                total_tested = summary_info.get('tested_students', 0)
                if total_tested > 0:
                    excellent = summary_info.get('excellent_count', 0)
                    good = summary_info.get('good_count', 0)
                    pass_count = summary_info.get('pass_count', 0)
                    fail = summary_info.get('fail_count', 0)
                    
                    # 估算平均分：优秀95分，良好85分，及格70分，不及格50分
                    total_score = excellent * 95 + good * 85 + pass_count * 70 + fail * 50
                    summary_info['average_score'] = total_score / total_tested if total_tested > 0 else 0
                else:
                    summary_info['average_score'] = 0
                
                # 提取班级序号用于排序
                num_match = re.search(r'([一二三四五六])(\d+)班', class_name)
                if num_match:
                    summary_info['class_num'] = int(num_match.group(2))
                else:
                    summary_info['class_num'] = 999
                
                class_summary_data[class_name] = summary_info
                
            except Exception as e:
                print(f"读取班级汇总表 {file_path} 时出错: {e}")
        
        return class_summary_data
    
    def calculate_rankings(self, class_data):
        """计算各项排名"""
        classes = list(class_data.keys())
        
        # 排名字典
        rankings = {}
        
        def assign_tied_rankings(sorted_data):
            """为相同值分配相同排名，并正确跳过后续排名"""
            rankings_dict = {}
            current_rank = 1
            
            for i, (cls, value) in enumerate(sorted_data):
                if i > 0 and value != sorted_data[i-1][1]:
                    # 值不同时，排名应该是当前索引+1
                    current_rank = i + 1
                rankings_dict[cls] = current_rank
            
            return rankings_dict
        
        # 优良率排名（优秀+良好）
        excellent_good_rates = []
        for cls in classes:
            total_students = class_data[cls].get('total_students', 0)
            if total_students > 0:
                excellent_good_count = class_data[cls].get('excellent_count', 0) + class_data[cls].get('good_count', 0)
                excellent_good_rate = excellent_good_count / total_students * 100
            else:
                excellent_good_rate = 0
            excellent_good_rates.append((cls, excellent_good_rate))
        
        excellent_good_rates.sort(key=lambda x: x[1], reverse=True)
        excellent_rankings = assign_tied_rankings(excellent_good_rates)
        for cls in classes:
            if cls not in rankings:
                rankings[cls] = {}
            rankings[cls]['excellent_rank'] = excellent_rankings[cls]
        
        # 及格率排名（及格+良好+优秀）
        for cls in classes:
            total = class_data[cls].get('tested_students', 0)
            if total > 0:
                pass_count = (class_data[cls].get('excellent_count', 0) + 
                            class_data[cls].get('good_count', 0) + 
                            class_data[cls].get('pass_count', 0))
                pass_rate = pass_count / total * 100
            else:
                pass_rate = 0
            class_data[cls]['pass_rate'] = pass_rate
        
        pass_rates = [(cls, class_data[cls].get('pass_rate', 0)) for cls in classes]
        pass_rates.sort(key=lambda x: x[1], reverse=True)
        pass_rankings = assign_tied_rankings(pass_rates)
        for cls in classes:
            rankings[cls]['pass_rank'] = pass_rankings[cls]
        
        # 平均分排名
        avg_scores = [(cls, class_data[cls].get('average_score', 0)) for cls in classes]
        avg_scores.sort(key=lambda x: x[1], reverse=True)
        score_rankings = assign_tied_rankings(avg_scores)
        for cls in classes:
            rankings[cls]['score_rank'] = score_rankings[cls]
        
        # 总排名（排名汇总：优良率排名 + 及格率排名 + 平均分排名，总和越小排名越高）
        total_rank_scores = []
        for cls in classes:
            # 直接把三个排名数字相加
            rank_sum = rankings[cls]['excellent_rank'] + rankings[cls]['pass_rank'] + rankings[cls]['score_rank']
            total_rank_scores.append((cls, rank_sum))
        
        # 按排名总和从小到大排序（总和越小排名越高）
        total_rank_scores.sort(key=lambda x: x[1])
        total_rankings = assign_tied_rankings(total_rank_scores)
        for cls in classes:
            rankings[cls]['total_rank'] = total_rankings[cls]
        
        return rankings
    
    def create_ranking_sheet(self, wb, grade, class_data, rankings):
        """创建年级排名表"""
        ws = wb.create_sheet(title=grade)
        
        # 创建边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 表头
        headers = ['项目', '班级名称', '参测人数', '未测人数', '优良率', '排名', '及格率', '排名', '平均分', '排名', '排名汇总', '总排名']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.border = thin_border
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')


       
        # 按总排名排序（排名越小越靠前）
        sorted_classes = sorted(class_data.keys(), key=lambda x: rankings[x].get('total_rank', 999))
        
        row = 2
        for class_name in sorted_classes:
            data = class_data[class_name]
            rank = rankings[class_name]
            
            # 计算未测人数
            total = data.get('total_students', 0)
            tested = data.get('tested_students', 0)
            untested = total - tested
            
            # 优良率 = (优秀 + 良好) / 班级总人数（应测人数）
            total_students = data.get('total_students', 0)
            if total_students > 0:
                excellent_good_count = data.get('excellent_count', 0) + data.get('good_count', 0)
                excellent_good_rate = excellent_good_count / total_students * 100
            else:
                excellent_good_rate = 0
            
            # 填充数据
            ws.cell(row=row, column=1, value=f"{grade}")
            ws.cell(row=row, column=2, value=class_name)
            ws.cell(row=row, column=3, value=tested)
            ws.cell(row=row, column=4, value=untested)
            ws.cell(row=row, column=5, value=f"{excellent_good_rate:.2f}%")
            ws.cell(row=row, column=6, value=rank.get('excellent_rank', 0))
            ws.cell(row=row, column=7, value=f"{data.get('pass_rate', 0):.2f}%")
            ws.cell(row=row, column=8, value=rank.get('pass_rank', 0))
            ws.cell(row=row, column=9, value=f"{data.get('average_score', 0):.2f}")
            ws.cell(row=row, column=10, value=rank.get('score_rank', 0))
            # 排名汇总：三个排名的总和
            rank_sum = rank.get('excellent_rank', 0) + rank.get('pass_rank', 0) + rank.get('score_rank', 0)
            ws.cell(row=row, column=11, value=rank_sum)
            ws.cell(row=row, column=12, value=rank.get('total_rank', 0))
            
            # 应用边框
            for col in range(1, 13):
                ws.cell(row=row, column=col).border = thin_border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
            
            row += 1
        
        # 设置列宽
        column_widths = [8, 12, 10, 10, 8, 6, 8, 6, 8, 6, 8, 8]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

         # === 新增：设置 F, G, J, K, M 列为浅蓝色背景（雾蓝色） ===
        light_blue_fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")  # 雾蓝色

        # 要上色的列（列字母）
        blue_cols = ['E', 'F', 'I', 'J', 'l']

        for col_letter in blue_cols:
            # 设置整列从第2行开始到底部的背景色
            for row in range(1, ws.max_row+1):
                cell = ws[f"{col_letter}{row}"]
                cell.fill = light_blue_fill
        
    
    def generate_ranking_report(self):
        """生成班级排名报告"""
        print("=== 开始生成班级排名报告 ===")
        
        # 加载年级数据（获取班级总人数）
        grade_class_info = self.load_grade_data()
    
        # 创建工作簿
        wb = Workbook()
        # 先不删除默认工作表，等有新工作表创建后再删除
        
        # 年级顺序 - 只处理四年级和六年级
        grade_order = ['四年级', '六年级']
        available_grades = [grade for grade in grade_order if grade in grade_class_info]
        
        sheets_created = 0  # 记录创建的工作表数量
        
        for grade in available_grades:
            print(f"\n=== 处理{grade} ===")
            
            # 加载班级汇总数据（优秀、良好、及格、不及格人数）
            class_data = self.load_class_summary_data(grade)
            if not class_data:
                print(f"{grade}没有找到班级汇总数据")
                continue
            
            # 注入 total_students（班级总人数）
            if grade in grade_class_info:
                for class_name, info in grade_class_info[grade].items():
                    if class_name in class_data:
                        class_data[class_name]['total_students'] = info['total_students']
            
            # 计算排名
            rankings = self.calculate_rankings(class_data)
            
            # 创建排名表
            self.create_ranking_sheet(wb, grade, class_data, rankings)
            sheets_created += 1
            
            print(f"{grade}排名表创建完成，共{len(class_data)}个班级")
        
        # 删除默认工作表（如果有新工作表创建）
        if sheets_created > 0 and 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # 保存文件
        try:
            wb.save(self.output_file)
            print(f"\n=== 报告生成完成 ===")
            print(f"保存位置: {self.output_file}")
        except PermissionError:
            alt_filename = f"{self.output_dir}\班级排名统计表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(alt_filename)
            print(f"保存位置: {alt_filename}")

def main():
    generator = ClassRankingGenerator()
    generator.generate_ranking_report()

if __name__ == "__main__":
    main()
